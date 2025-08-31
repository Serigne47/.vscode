# utils/report_generator.py
"""
Générateur de rapports pour l'analyse d'appels d'offres
Formats supportés: Excel, Word, PDF
"""
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
import json

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, PieChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("⚠️ openpyxl/pandas non installé. Export Excel désactivé.")

try:
    from docx import Document as WordDocument
    from docx.shared import Inches, Pt, RGBColor
    from docx.enum.text import WD_ALIGN_PARAGRAPH
    from docx.enum.style import WD_STYLE_TYPE
    WORD_AVAILABLE = True
except ImportError:
    WORD_AVAILABLE = False
    logging.warning("⚠️ python-docx non installé. Export Word désactivé.")

logger = logging.getLogger(__name__)

class AOReportGenerator:
    """
    Générateur de rapports multi-format pour l'analyse AO
    """
    
    def __init__(self):
        """
        Initialise le générateur de rapports
        """
        self.excel_available = EXCEL_AVAILABLE
        self.word_available = WORD_AVAILABLE
        
        # Styles Excel
        self.excel_styles = {
            'header': {
                'font': Font(bold=True, size=14, color="FFFFFF"),
                'fill': PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'subheader': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
                'alignment': Alignment(horizontal="left", vertical="center")
            },
            'critical': {
                'font': Font(bold=True, color="FF0000"),
                'fill': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            },
            'warning': {
                'font': Font(bold=True, color="FF9900"),
                'fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            },
            'success': {
                'font': Font(bold=True, color="008000"),
                'fill': PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
            }
        }
        
        logger.info(f"✅ AOReportGenerator initialisé (Excel: {self.excel_available}, Word: {self.word_available})")
    
    def generate_analysis_report(
        self,
        analysis_results: Dict[str, Any],
        output_path: Path,
        formats: List[str] = ["excel", "word", "json"]
    ) -> Dict[str, Path]:
        """
        Génère les rapports dans les formats demandés
        
        Args:
            analysis_results: Résultats de l'analyse
            output_path: Dossier de sortie
            formats: Formats à générer
            
        Returns:
            Chemins des fichiers générés
        """
        output_path = Path(output_path)
        output_path.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"ao_analysis_{timestamp}"
        
        generated_files = {}
        
        # JSON (toujours disponible)
        if "json" in formats:
            json_path = output_path / f"{base_name}.json"
            self._export_json(analysis_results, json_path)
            generated_files["json"] = json_path
        
        # Excel
        if "excel" in formats and self.excel_available:
            excel_path = output_path / f"{base_name}.xlsx"
            self.export_to_excel(analysis_results, excel_path)
            generated_files["excel"] = excel_path
        
        # Word
        if "word" in formats and self.word_available:
            word_path = output_path / f"{base_name}.docx"
            self.export_to_word(analysis_results, word_path)
            generated_files["word"] = word_path
        
        logger.info(f"✅ Rapports générés: {list(generated_files.keys())}")
        return generated_files
    
    def _export_json(self, data: Dict, path: Path):
        """
        Export JSON avec formatage
        """
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        logger.info(f"📄 JSON exporté: {path}")
    
    def export_to_excel(self, report: Dict[str, Any], output_path: Path):
        """
        Génère un rapport Excel structuré
        """
        if not self.excel_available:
            logger.error("Export Excel non disponible (installez openpyxl et pandas)")
            return
        
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # 1. Résumé Exécutif
        self._create_executive_summary_sheet(wb, report.get('executive_summary', {}))
        
        # 2. Identité AO
        self._create_identity_sheet(wb, report.get('fiche_identite', {}))
        
        # 3. Volumes
        self._create_volumes_sheet(wb, report.get('analyse_volumes', {}))
        
        # 4. Aspects Financiers
        self._create_financial_sheet(wb, report.get('conditions_financieres', {}))
        
        # 5. Aspects Juridiques
        self._create_legal_sheet(wb, report.get('aspects_juridiques', {}))
        
        # 6. Exigences Opérationnelles
        self._create_operational_sheet(wb, report.get('exigences_operationnelles', {}))
        
        # 7. Timeline
        self._create_timeline_sheet(wb, report.get('calendrier', {}))
        
        # 8. Risques et Recommandations
        self._create_risks_sheet(wb, report)
        
        # Sauvegarder
        wb.save(output_path)
        logger.info(f"📊 Excel exporté: {output_path}")
    
    def _create_executive_summary_sheet(self, wb, summary_data: Dict):
        """
        Crée la feuille de résumé exécutif
        """
        ws = wb.create_sheet("Résumé Exécutif")
        
        # Titre
        ws['A1'] = "RÉSUMÉ EXÉCUTIF - ANALYSE APPEL D'OFFRES"
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], self.excel_styles['header'])
        
        # Informations clés
        row = 3
        key_info = [
            ("Client", summary_data.get('client', 'N/A')),
            ("Référence AO", summary_data.get('reference', 'N/A')),
            ("Date limite", summary_data.get('date_limite', 'N/A')),
            ("Volumes estimés", summary_data.get('volume_estime', 'N/A')),
            ("Durée contrat", f"{summary_data.get('duree_contrat', 'N/A')} mois"),
            ("Complexité", summary_data.get('complexite', 'N/A')),
            ("Risque global", summary_data.get('risque_global', 'N/A')),
            ("Complétude données", summary_data.get('completude_donnees', 'N/A'))
        ]
        
        for label, value in key_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Recommandation Go/No-Go
        row += 1
        go_no_go = summary_data.get('go_no_go_recommendation', {})
        ws[f'A{row}'] = "DÉCISION RECOMMANDÉE"
        ws.merge_cells(f'A{row}:B{row}')
        self._apply_style(ws[f'A{row}'], self.excel_styles['subheader'])
        
        row += 1
        ws[f'A{row}'] = go_no_go.get('decision', 'N/A')
        ws[f'B{row}'] = f"Confiance: {go_no_go.get('confidence', 'N/A')}"
        
        # Appliquer le style selon la décision
        if 'GO' in str(go_no_go.get('decision', '')):
            if 'CONDITIONNEL' in str(go_no_go.get('decision', '')):
                self._apply_style(ws[f'A{row}'], self.excel_styles['warning'])
            else:
                self._apply_style(ws[f'A{row}'], self.excel_styles['success'])
        else:
            self._apply_style(ws[f'A{row}'], self.excel_styles['critical'])
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    def _create_identity_sheet(self, wb, identity_data: Dict):
        """
        Crée la feuille d'identité de l'AO
        """
        ws = wb.create_sheet("Identité AO")
        
        ws['A1'] = "IDENTIFICATION DE L'APPEL D'OFFRES"
        ws.merge_cells('A1:D1')
        self._apply_style(ws['A1'], self.excel_styles['header'])
        
        # Client
        row = 3
        ws[f'A{row}'] = "CLIENT"
        self._apply_style(ws[f'A{row}'], self.excel_styles['subheader'])
        row += 1
        
        client = identity_data.get('client', {})
        client_info = [
            ("Nom", client.get('nom', 'N/A')),
            ("Groupe", client.get('groupe', 'N/A')),
            ("Division", client.get('division', 'N/A'))
        ]
        
        for label, value in client_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            row += 1
        
        # Type de marché
        row += 1
        ws[f'A{row}'] = "TYPE DE MARCHÉ"
        self._apply_style(ws[f'A{row}'], self.excel_styles['subheader'])
        row += 1
        
        marche = identity_data.get('type_marche', {})
        ws[f'A{row}'] = "Nature"
        ws[f'B{row}'] = marche.get('nature', 'N/A')
        row += 1
        ws[f'A{row}'] = "Attribution"
        ws[f'B{row}'] = marche.get('attribution', 'N/A')
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _create_volumes_sheet(self, wb, volumes_data: Dict):
        """
        Crée la feuille des volumes
        """
        ws = wb.create_sheet("Volumes")
        
        ws['A1'] = "ANALYSE DES VOLUMES"
        ws.merge_cells('A1:E1')
        self._apply_style(ws['A1'], self.excel_styles['header'])
        
        # Agrégats
        row = 3
        ws[f'A{row}'] = "VOLUMES AGRÉGÉS"
        self._apply_style(ws[f'A{row}'], self.excel_styles['subheader'])
        row += 1
        
        aggregats = volumes_data.get('aggregats', {})
        volume_info = [
            ("TEU annuel", aggregats.get('total_teu_annuel', 0)),
            ("Tonnage annuel", aggregats.get('total_tonnage_annuel', 0)),
            ("M³ annuel", aggregats.get('total_m3_annuel', 0)),
            ("Palettes annuel", aggregats.get('total_palettes_annuel', 0))
        ]
        
        for label, value in volume_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = f"{value:,.0f}" if isinstance(value, (int, float)) else str(value)
            row += 1
        
        # Détail des volumes si disponibles
        detail = volumes_data.get('detail_volumes', {})
        if detail.get('conteneurs'):
            row += 2
            ws[f'A{row}'] = "DÉTAIL CONTENEURS"
            self._apply_style(ws[f'A{row}'], self.excel_styles['subheader'])
            row += 1
            
            # Créer un tableau
            headers = ["Type", "Valeur", "Unité", "Période", "Nature"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            
            row += 1
            for item in detail['conteneurs'][:10]:  # Limiter à 10 lignes
                ws.cell(row=row, column=1, value=item.get('type', ''))
                ws.cell(row=row, column=2, value=item.get('valeur', ''))
                ws.cell(row=row, column=3, value=item.get('unite', ''))
                ws.cell(row=row, column=4, value=item.get('periode', ''))
                ws.cell(row=row, column=5, value=item.get('nature', ''))
                row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_financial_sheet(self, wb, financial_data: Dict):
        """
        Crée la feuille des aspects financiers
        """
        ws = wb.create_sheet("Financier")
        
        ws['A1'] = "CONDITIONS FINANCIÈRES"
        ws.merge_cells('A1:D1')
        self._apply_style(ws['A1'], self.excel_styles['header'])
        
        row = 3
        
        # Conditions de paiement
        ws[f'A{row}'] = "CONDITIONS DE PAIEMENT"
        self._apply_style(ws[f'A{row}'], self.excel_styles['subheader'])
        row += 1
        
        paiement = financial_data.get('conditions_paiement', {})
        ws[f'A{row}'] = "Délai"
        ws[f'B{row}'] = f"{paiement.get('delai_jours', 'N/A')} jours"
        row += 1
        ws[f'A{row}'] = "Base calcul"
        ws[f'B{row}'] = paiement.get('base_calcul', 'N/A')
        
        # Devise
        row += 2
        ws[f'A{row}'] = "DEVISE"
        self._apply_style(ws[f'A{row}'], self.excel_styles['subheader'])
        row += 1
        
        devise = financial_data.get('devise', {})
        ws[f'A{row}'] = "Facturation"
        ws[f'B{row}'] = devise.get('facturation', 'N/A')
        row += 1
        ws[f'A{row}'] = "Paiement"
        ws[f'B{row}'] = devise.get('paiement', 'N/A')
        
        # Points d'attention
        points = financial_data.get('points_attention', [])
        if points:
            row += 2
            ws[f'A{row}'] = "POINTS D'ATTENTION"
            self._apply_style(ws[f'A{row}'], self.excel_styles['warning'])
            row += 1
            for point in points[:5]:
                ws[f'A{row}'] = "⚠"
                ws[f'B{row}'] = point
                row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def _create_legal_sheet(self, wb, legal_data: Dict):
        """
        Crée la feuille des aspects juridiques
        """
        ws = wb.create_sheet("Juridique")
        
        ws['A1'] = "ANALYSE JURIDIQUE"
        ws.merge_cells('A1:D1')
        self._apply_style(ws['A1'], self.excel_styles['header'])
        
        row = 3
        
        # Score de protection
        protection = legal_data.get('score_protection', {})
        ws[f'A{row}'] = "SCORE DE PROTECTION"
        ws[f'B{row}'] = f"{protection.get('score', 0)}/100"
        ws[f'C{row}'] = protection.get('niveau', 'N/A')
        
        if protection.get('score', 0) >= 70:
            self._apply_style(ws[f'B{row}'], self.excel_styles['success'])
        elif protection.get('score', 0) >= 40:
            self._apply_style(ws[f'B{row}'], self.excel_styles['warning'])
        else:
            self._apply_style(ws[f'B{row}'], self.excel_styles['critical'])
        
        # Clauses manquantes
        missing = legal_data.get('clauses_manquantes', [])
        if missing:
            row += 2
            ws[f'A{row}'] = "CLAUSES MANQUANTES"
            self._apply_style(ws[f'A{row}'], self.excel_styles['critical'])
            row += 1
            for clause in missing[:10]:
                ws[f'A{row}'] = "❌"
                ws[f'B{row}'] = clause
                row += 1
        
        # Analyse des risques
        risk_analysis = legal_data.get('analyse_risques', {})
        if risk_analysis:
            row += 2
            ws[f'A{row}'] = "ANALYSE DES RISQUES JURIDIQUES"
            self._apply_style(ws[f'A{row}'], self.excel_styles['subheader'])
            row += 1
            
            ws[f'A{row}'] = "Niveau de risque"
            ws[f'B{row}'] = risk_analysis.get('level', 'N/A')
            
            # Style selon le niveau
            if risk_analysis.get('level') in ['CRITIQUE', 'ÉLEVÉ']:
                self._apply_style(ws[f'B{row}'], self.excel_styles['critical'])
            elif risk_analysis.get('level') == 'MODÉRÉ':
                self._apply_style(ws[f'B{row}'], self.excel_styles['warning'])
            else:
                self._apply_style(ws[f'B{row}'], self.excel_styles['success'])
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20